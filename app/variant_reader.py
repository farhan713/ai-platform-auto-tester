"""Read a variant-questions workbook.

Format (see Variant-Questions-Template.xlsx):
    Column A holds questions. The ORIGINAL question of a group is BOLD; the
    rows below it (not bold) are its 3-5 variants, until the next bold row.

We flatten every question into a single ordered list (so the existing SQL
Agent runner can execute them sequentially) AND return a group map describing
which qid is the original and which qids are its variants. Sheets named like
'instructions' / 'how to use' / 'readme' are ignored.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_SKIP_SHEET_HINTS = ("instruction", "how to use", "how-to", "readme", "read me", "guide")
_HEADER_HINTS = {
    "question", "questions", "natural_language_query", "nl_query", "query",
}


def _is_bold(cell) -> bool:
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


def read_variant_questions(xlsx_path: str | Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Return (questions, groups).

    questions: [{id, natural_language_query, expected_sql}] in run order.
    groups:    [{group_id, sheet, original_qid, variant_qids: [...],
                 original_text}]  — one entry per original question.

    Rules:
      * A bold cell in column A starts a new group (it is the original).
      * Non-bold rows after it are variants of the current group.
      * If a sheet has NO bold cells at all, we fall back to treating the
        first row of the sheet as the original and the rest as variants
        (so a user who forgot to bold still gets a single group, rather
        than every row being its own original).
    """
    p = Path(xlsx_path)
    if not p.exists():
        raise FileNotFoundError(f"variant questions file not found: {p}")

    # Need formatting (bold), so NOT read_only.
    wb = load_workbook(filename=str(p), data_only=True)

    questions: list[dict[str, Any]] = []
    groups: list[dict[str, Any]] = []
    next_qid = 1
    next_gid = 1

    for sheet_name in wb.sheetnames:
        if any(h in sheet_name.lower() for h in _SKIP_SHEET_HINTS):
            continue
        ws = wb[sheet_name]

        # Collect non-empty column-A cells with their bold flag.
        cells: list[tuple[str, bool]] = []
        for idx, row in enumerate(ws.iter_rows(min_col=1, max_col=1), start=1):
            cell = row[0]
            val = "" if cell.value is None else str(cell.value).strip()
            if not val:
                continue
            # Skip an obvious header in the first populated row.
            if not cells and val.lower() in _HEADER_HINTS:
                continue
            # Also skip the template's instructional header text.
            if not cells and val.lower().startswith("question ("):
                continue
            cells.append((val, _is_bold(cell)))

        if not cells:
            continue

        sheet_has_bold = any(b for _, b in cells)

        cur_group: dict[str, Any] | None = None
        for i, (text, bold) in enumerate(cells):
            is_original = bold if sheet_has_bold else (i == 0)
            qid = next_qid
            next_qid += 1
            questions.append({
                "id": qid,
                "natural_language_query": text,
                "expected_sql": "",
            })
            if is_original or cur_group is None:
                cur_group = {
                    "group_id": next_gid,
                    "sheet": sheet_name,
                    "original_qid": qid,
                    "original_text": text,
                    "variant_qids": [],
                }
                next_gid += 1
                groups.append(cur_group)
            else:
                cur_group["variant_qids"].append(qid)

    if not questions:
        raise ValueError(f"no usable questions found in {p}")
    return questions, groups
